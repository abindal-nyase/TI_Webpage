"""
Project TI Classifier and Geocoder
====================================
Reads the 'Selected' tab, columns B:J from the BD tracking workbook and:
  1. Classifies each project as TI / NOT_TI / AMBIGUOUS using keyword rules
     (optionally resolves AMBIGUOUS rows via Claude Haiku)
  2. Geocodes each project (Mapbox default) -- skip with --classify-only

Outputs a styled Excel file with classification results and lat/lng.

Requirements:
    pip install pandas openpyxl anthropic requests

Usage:
    # Classification only (no geocoding, no API key needed for rules)
    python project_ti_classifier.py --classify-only

    # Full run with Mapbox geocoding
    python project_ti_classifier.py --mapbox-key YOUR_TOKEN

    # Include Claude Haiku for ambiguous rows
    python project_ti_classifier.py --mapbox-key TOKEN --anthropic-key KEY

    # Dry run: see ambiguous row count and cost estimate only
    python project_ti_classifier.py --dry-run
"""

import argparse
import json
import logging
import os
import re
import time
import urllib.parse

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

load_dotenv()
AI_KEY = os.getenv("ANTHROPIC_API_KEY")
MAPBOX_KEY = os.getenv("MAPBOX_TOKEN")

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
log = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Actual column names from 'Selected' tab, columns B:J
# ---------------------------------------------------------------------------
COL_NUMBER       = "PROJECT NUMBER"
COL_NAME         = "PROJECT NAME"
COL_ADDRESS      = "PROJECT ADDRESS"
COL_CITY         = "CITY"
COL_STATE        = "STATE"
COL_ZIP          = "ZIP CODE"
COL_DESCRIPTION  = "Description"
DESC_CHAR_LIMIT  = 500   # truncate descriptions before sending to AI


# ---------------------------------------------------------------------------
# Rule-based triage
# ---------------------------------------------------------------------------
_STRONG_TI = re.compile(
    r"\bT\.?I\.?\b"                    # TI, T.I., T.I  (word-boundary prevents Tier)
    r"|\bTenant\s+Improvement\b"
    r"|\bTenant\b"
    r"|\bFit[-\s]?[Oo]ut\b"
    r"|\bInterior\s+Improvement\b"
    r"|\bOffice\s+(Renovation|Remodel)\b"
    r"|\bRemodel\b"
    r"|\bRefurbi?sh\b"
    r"|\bRenovation\b"
    r"|\bRepair\b"
    r"|\bRevision\b"
    r"|\bUpgrade\b"
    r"|\bAnchorage\b"
    r"|\bAddition\b",
    re.IGNORECASE,
)

_STRONG_NOT_TI = re.compile(
    r"\bNew\s+Construction\b"
    r"|\bGround[-\s]?[Uu]p\b"
    r"|\bMaster\s+Plan\b"
    r"|\bInfrastructure\b"
    r"|\b(Civil|Bridge|Highway|Road)\b"
    r"|\bParking\s+(Lot|Garage|Structure)\b"
    r"|\bWastewater\b"
    r"|\bWater\s+Treatment\b"
    r"|\bSubstation\b"
    r"|\bFeasibility\s+(Study|Report)\b"
    r"|\bEvaluation\b"
    r"|\bStudy\b"
    r"|\bAnalysis\b",
    re.IGNORECASE,
)


def _apply_rules(text: str) -> str:
    """Apply rules to a cell to determine if it is TI, NOT_TI, or AMBIGUOUS"""
    if _STRONG_TI.search(text):
        return "TI"
    if _STRONG_NOT_TI.search(text):
        return "NOT_TI"
    return "AMBIGUOUS"


def rule_classify(name: str, desc: str = "") -> str:
    """Check name first; fall back to description if name is AMBIGUOUS."""
    result = _apply_rules(name)
    if result == "AMBIGUOUS" and desc:
        result = _apply_rules(desc)
    return result


# ---------------------------------------------------------------------------
# Claude Haiku classification for ambiguous rows
# ---------------------------------------------------------------------------
SYSTEM_PROMPT = (
    "You are a construction project classifier. "
    "A Tenant Improvement (TI) is any modification to an existing building interior "
    "for a tenant: fit-outs, reconfigs, remodels, repairs, upgrades of leased space. "
    "NOT TI: new ground-up construction, pure civil/infrastructure, feasibility studies, "
    "evaluations, studies, analyses. "
    "Reply ONLY with a valid JSON array. No markdown, no text outside the array."
)

RESPONSE_SCHEMA = {
    "type": "array",
    "items": {
        "type": "object",
        "required": ["id", "is_ti"],
        "properties": {
            "id":    {"type": "integer"},
            "is_ti": {"type": "boolean"},
        },
        "additionalProperties": False,
    },
}


def ai_classify_batch(rows: list, _retry: bool = True) -> list:
    import anthropic
    client = anthropic.Anthropic(api_key=AI_KEY)
    payload = json.dumps(
        [{"id": i, "name": r["name"], "description": r.get("desc", "")[:DESC_CHAR_LIMIT]}
         for i, r in enumerate(rows)],
        ensure_ascii=False,
    )
    prompt = (
        "Classify each project. Return ONLY a JSON array matching this schema:\n"
        + json.dumps(RESPONSE_SCHEMA, indent=2)
        + "\n\nProjects:\n"
        + payload
    )
    try:
        resp = client.messages.create(
            model="claude-haiku-4-5-20251001",
            max_tokens=8192,
            system=SYSTEM_PROMPT,
            messages=[{"role": "user", "content": prompt}],
        )
        raw = resp.content[0].text.strip()
        raw = re.sub(r"^```[a-z]*\n?|\n?```$", "", raw).strip()
        by_id = {r["id"]: r for r in json.loads(raw)}
        return [
            by_id.get(i, {"id": i, "is_ti": False})
            for i in range(len(rows))
        ]
    except anthropic.RateLimitError as exc:
        if _retry:
            log.warning(f"  Rate limit hit — waiting 60s before retry: {exc}")
            time.sleep(60)
            return ai_classify_batch(rows, _retry=False)
        log.warning(f"  Rate limit retry also failed: {exc}")
        return [{"id": i, "is_ti": False} for i in range(len(rows))]
    except Exception as exc:
        log.warning(f"  AI batch error: {exc}")
        return [{"id": i, "is_ti": False} for i in range(len(rows))]


def classify_all(names, descs, use_ai, batch_size, dry_run):
    """Returns parallel lists: (labels, methods)."""
    n      = len(names)
    tiers  = [rule_classify(names[i], descs[i]) for i in range(n)]
    amb    = [i for i, t in enumerate(tiers) if t == "AMBIGUOUS"]
    pct    = len(amb) / n * 100 if n else 0.0

    log.info(
        f"Rule triage -> TI: {tiers.count('TI'):,}  |  "
        f"NOT_TI: {tiers.count('NOT_TI'):,}  |  "
        f"AMBIGUOUS: {len(amb):,} ({pct:.1f}%)"
    )
    tok_in  = len(amb) * 50
    tok_out = len(amb) * 10
    est     = tok_in * 0.0000008 + tok_out * 0.000004
    log.info(
        f"  Estimated Haiku cost: ${est:.4f}  "
        f"({tok_in:,} input / {tok_out:,} output tokens)"
    )

    labels  = ["YES" if t == "TI" else "NO" if t == "NOT_TI" else "AMBIGUOUS" for t in tiers]
    methods = ["rule"] * n

    if dry_run:
        log.info("  --dry-run: skipping all API calls.")
        return labels, methods

    if not use_ai or not amb:
        return labels, methods

    # Resolve cached results first
    ai_cache   = _load_ai_cache()
    cache_hits = 0
    uncached   = []
    for idx in amb:
        key = _ai_cache_key(names[idx], descs[idx])
        if key in ai_cache:
            labels[idx]  = "YES" if ai_cache[key] else "NO"
            methods[idx] = "AI-Haiku"
            cache_hits  += 1
        else:
            uncached.append(idx)

    if cache_hits:
        log.info(f"  AI cache hits: {cache_hits:,} / {len(amb):,}")
    if not uncached:
        return labels, methods

    log.info(f"  Sending {len(uncached):,} uncached rows to Haiku (batch size {batch_size}) ...")
    payload_rows = [{"name": names[i], "desc": descs[i]} for i in uncached]

    processed  = 0
    last_save  = 0
    last_delay = 0

    for start in range(0, len(payload_rows), batch_size):
        end     = min(start + batch_size, len(payload_rows))
        log.info(f"    Batch {start + 1}-{end} / {len(payload_rows)}")
        results = ai_classify_batch(payload_rows[start:end])

        for local_i, r in enumerate(results):
            idx   = uncached[start + local_i]
            is_ti = r.get("is_ti", False)
            labels[idx]  = "YES" if is_ti else "NO"
            methods[idx] = "AI-Haiku"
            ai_cache[_ai_cache_key(names[idx], descs[idx])] = is_ti

        processed += end - start

        if processed - last_save >= 500:
            _save_ai_cache(ai_cache)
            log.info(f"  AI cache saved ({processed:,} records processed)")
            last_save = processed

        if processed - last_delay >= 1000:
            log.info(f"  Pausing 5s after {processed:,} records ...")
            time.sleep(5)
            last_delay = processed

    _save_ai_cache(ai_cache)
    return labels, methods


# ---------------------------------------------------------------------------
# AI classification cache (persists across interrupted runs)
# ---------------------------------------------------------------------------
AI_CACHE_FILE = "ai_classify_cache.json"


def _load_ai_cache():
    if os.path.exists(AI_CACHE_FILE):
        with open(AI_CACHE_FILE) as f:
            return json.load(f)
    return {}


def _save_ai_cache(cache):
    with open(AI_CACHE_FILE, "w") as f:
        json.dump(cache, f, indent=2)


def _ai_cache_key(name, desc):
    return f"{name}|||{desc[:DESC_CHAR_LIMIT]}"


# ---------------------------------------------------------------------------
# Geocoding (Mapbox default, with local cache)
# ---------------------------------------------------------------------------
CACHE_FILE = "geocode_cache.json"


def _load_cache():
    if os.path.exists(CACHE_FILE):
        with open(CACHE_FILE) as f:
            return json.load(f)
    return {}


def _save_cache(cache):
    with open(CACHE_FILE, "w") as f:
        json.dump(cache, f, indent=2)


def _build_address(row):
    parts = []
    for col in (COL_ADDRESS, COL_CITY, COL_STATE, COL_ZIP):
        if col in row.index:
            v = str(row[col]).strip().replace("\n", " ")
            if v and v.lower() not in ("nan", "none", ""):
                parts.append(v)
    return ", ".join(parts)


def _mapbox_geocode(addr):
    import requests
    resp = requests.get(
        f"https://api.mapbox.com/geocoding/v5/mapbox.places/{urllib.parse.quote(addr)}.json",
        params={"access_token": MAPBOX_KEY, "limit": 1, "country": "us"},
        timeout=10,
    )
    data = resp.json()
    features = data.get("features", [])
    if features:
        lng, lat = features[0]["geometry"]["coordinates"]
        return lat, lng, features[0].get("place_name", "")
    return None, None, "not found"


def _nominatim_geocode(addr, geolocator):
    from geopy.exc import GeocoderServiceError, GeocoderTimedOut
    try:
        loc = geolocator.geocode(addr, timeout=10)
        time.sleep(1.05)
        if loc:
            return loc.latitude, loc.longitude, loc.address
        parts = [p.strip() for p in addr.split(",") if p.strip()]
        if len(parts) > 2:
            loc = geolocator.geocode(", ".join(parts[-3:]), timeout=10)
            time.sleep(1.05)
            if loc:
                return loc.latitude, loc.longitude, f"(partial) {loc.address}"
        return None, None, "not found"
    except (GeocoderTimedOut, GeocoderServiceError) as exc:
        return None, None, str(exc)


def geocode_all(df, geocoder):
    cache = _load_cache()
    lats, lngs, fmts = [], [], []

    geolocator = None
    if geocoder == "nominatim":
        from geopy.geocoders import Nominatim
        geolocator = Nominatim(user_agent="project_ti_classifier/1.0")

    total = len(df)
    hits  = 0
    for i, (_, row) in enumerate(df.iterrows()):
        addr = _build_address(row)
        if not addr:
            addr = f"{str(row.get(COL_NAME, ''))} Los Angeles, CA"
        cache_key = addr.lower().strip()

        if cache_key in cache:
            e = cache[cache_key]
            lats.append(e["lat"])
            lngs.append(e["lng"])
            fmts.append(e["fmt"])
            hits += 1
            continue

        if geocoder == "mapbox":
            lat, lng, fmt = _mapbox_geocode(addr)
        elif geocoder == "nominatim":
            lat, lng, fmt = _nominatim_geocode(addr, geolocator)
        else:
            lat, lng, fmt = None, None, "no geocoder configured"

        cache[cache_key] = {"lat": lat, "lng": lng, "fmt": fmt}
        lats.append(lat)
        lngs.append(lng)
        fmts.append(fmt)

        if (i + 1) % 100 == 0:
            _save_cache(cache)
            log.info(f"  Geocoded {i + 1:,}/{total:,}  (cache hits: {hits:,})")

    _save_cache(cache)
    geo_ok = sum(1 for lat in lats if lat)
    log.info(f"  Done: {geo_ok:,}/{total:,} resolved  |  {hits:,} from cache")
    return lats, lngs, fmts


# ---------------------------------------------------------------------------
# Styled Excel output
# ---------------------------------------------------------------------------
def write_excel(df_out, path):
    df_out.to_excel(path, index=False, sheet_name="Projects")
    wb = load_workbook(path)
    ws = wb["Projects"]

    H_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=10)
    H_FILL   = PatternFill("solid", start_color="1F3864")
    TI_FILL  = PatternFill("solid", start_color="C6EFCE")   # green
    NO_FILL  = PatternFill("solid", start_color="FFEB9C")   # yellow
    AMB_FILL = PatternFill("solid", start_color="F4B183")   # orange
    ER_FILL  = PatternFill("solid", start_color="FFC7CE")   # red (geocode fail)
    AI_FILL  = PatternFill("solid", start_color="DDEBF7")   # blue (AI classified)
    thin     = Side(style="thin", color="CCCCCC")
    BDR      = Border(left=thin, right=thin, top=thin, bottom=thin)
    CTR      = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [c.value for c in ws[1]]

    def col_idx(name):
        return headers.index(name) + 1 if name in headers else None

    ti_ci   = col_idx("AUTO_TI")
    lat_ci  = col_idx("LATITUDE")
    meth_ci = col_idx("METHOD")

    for cell in ws[1]:
        cell.font = H_FONT
        cell.fill = H_FILL
        cell.alignment = CTR
        cell.border = BDR
    ws.row_dimensions[1].height = 26

    for rn in range(2, ws.max_row + 1):
        ws.row_dimensions[rn].height = 18
        for cell in ws[rn]:
            cell.border = BDR
            cell.alignment = CTR
        if ti_ci:
            c = ws.cell(rn, ti_ci)
            c.fill = (TI_FILL if c.value == "YES" else
                      AMB_FILL if c.value == "AMBIGUOUS" else NO_FILL)
        if lat_ci and not ws.cell(rn, lat_ci).value:
            ws.cell(rn, lat_ci).fill = ER_FILL
        if meth_ci and ws.cell(rn, meth_ci).value == "AI-Haiku":
            ws.cell(rn, meth_ci).fill = AI_FILL

    for cn, col in enumerate(ws.columns, 1):
        max_w = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(cn)].width = min(max(max_w + 2, 10), 55)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    # Summary tab
    ws2   = wb.create_sheet("Summary")
    total = len(df_out)
    ti_n  = (df_out["AUTO_TI"] == "YES").sum()
    no_n  = (df_out["AUTO_TI"] == "NO").sum()
    amb_n = (df_out["AUTO_TI"] == "AMBIGUOUS").sum()
    geo_n = df_out["LATITUDE"].notna().sum() if "LATITUDE" in df_out else 0
    ai_n  = (df_out["METHOD"] == "AI-Haiku").sum() if "METHOD" in df_out else 0

    summary_rows = [
        ["Metric",              "Value"],
        ["Total Projects",      total],
        ["TI (YES)",            int(ti_n)],
        ["Not TI (NO)",         int(no_n)],
        ["Ambiguous",           int(amb_n)],
        ["% TI",                f"{ti_n / total * 100:.1f}%" if total else "0%"],
        ["",                    ""],
        ["Rule-classified",     total - int(ai_n)],
        ["AI-Haiku classified", int(ai_n)],
        ["",                    ""],
        ["Geocoded OK",         int(geo_n)],
        ["Geocode Failed",      total - int(geo_n)],
    ]
    for r, (k, v) in enumerate(summary_rows, 1):
        for c, val in enumerate([k, v], 1):
            cell = ws2.cell(r, c, val)
            cell.border = BDR
            if r == 1:
                cell.font = H_FONT
                cell.fill = H_FILL
                cell.alignment = CTR
    ws2.column_dimensions["A"].width = 26
    ws2.column_dimensions["B"].width = 18

    wb.save(path)
    log.info(f"Saved -> {path}")


# ---------------------------------------------------------------------------
# Pipeline
# ---------------------------------------------------------------------------
def process(
    input_path,
    output_path,
    sheet="Selected",
    usecols="A:G",
    classify_only=False,
    use_ai=False,
    geocoder="mapbox",
    batch_size=50,
    dry_run=False,
):
    log.info(f"Reading: {input_path}  |  sheet='{sheet}'  |  cols={usecols}")
    df = pd.read_excel(input_path, sheet_name=sheet, usecols=usecols, dtype=str).fillna("")
    log.info(f"  {len(df):,} rows  |  columns: {list(df.columns)}")

    names = df[COL_NAME].tolist()
    descs = df[COL_DESCRIPTION].tolist()

    log.info("Step 1  TI Classification ...")
    labels, methods = classify_all(names, descs, use_ai, batch_size, dry_run)

    out = df.copy()
    out["AUTO_TI"] = labels    # YES / NO / AMBIGUOUS
    out["METHOD"]  = methods   # rule / AI-Haiku

    if not classify_only:
        log.info(f"Step 2  Geocoding ({geocoder}) ...")
        lats, lngs, geo_fmt = geocode_all(df, geocoder)
        out["LATITUDE"]         = lats
        out["LONGITUDE"]        = lngs
        out["GEOCODED_ADDRESS"] = geo_fmt
    else:
        out["LATITUDE"]         = None
        out["LONGITUDE"]        = None
        out["GEOCODED_ADDRESS"] = ""
        log.info("  Geocoding skipped (--classify-only)")

    write_excel(out, output_path)


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------
def main():
    p = argparse.ArgumentParser(description="TI Classifier and Geocoder")
    p.add_argument("--input",
                   default="Project-BD Number Tracking - MASTER 2026.xlsx")
    p.add_argument("--output",         default="output_classified.xlsx")
    p.add_argument("--sheet",          default="Selected")
    p.add_argument("--usecols",        default="A:G")
    p.add_argument("--classify-only",  action="store_true",
                   help="Skip geocoding, output classification only")
    p.add_argument("--geocoder",       default="mapbox",
                   choices=["mapbox", "nominatim"])
    p.add_argument("--no-ai",          action="store_true",
                   help="Rule-based only, no Haiku calls")
    p.add_argument("--batch-size",     type=int, default=50)
    p.add_argument("--dry-run",        action="store_true",
                   help="Estimate AI cost only, no API calls")
    a = p.parse_args()

    process(
        input_path    = a.input,
        output_path   = a.output,
        sheet         = a.sheet,
        usecols       = a.usecols,
        classify_only = a.classify_only,
        use_ai        = not a.no_ai,
        geocoder      = a.geocoder,
        batch_size    = a.batch_size,
        dry_run       = a.dry_run,
    )


if __name__ == "__main__":
    main()
